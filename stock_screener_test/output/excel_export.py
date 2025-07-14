"""
Excel report generation
"""

import pandas as pd
import numpy as np
from datetime import datetime
import logging
from typing import Dict, List
import os

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

class ExcelReporter:
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        
        if not EXCEL_AVAILABLE:
            self.logger.warning("openpyxl not available. Excel reports will be basic CSV files.")
    
    def generate_report(self, screening_results: Dict[str, pd.DataFrame], 
                       output_file: str, include_charts: bool = False):
        """Generate comprehensive Excel report"""
        
        if not EXCEL_AVAILABLE:
            self._generate_csv_reports(screening_results, output_file)
            return
        
        try:
            # Create workbook
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Add summary sheet
            self._create_summary_sheet(wb, screening_results)
            
            # Add individual strategy sheets
            for strategy_name, results_df in screening_results.items():
                if not results_df.empty:
                    self._create_strategy_sheet(wb, strategy_name, results_df)
            
            # Add analysis sheet
            self._create_analysis_sheet(wb, screening_results)
            
            # Save workbook
            wb.save(output_file)
            self.logger.info(f"Excel report saved to {output_file}")
            
        except Exception as e:
            self.logger.error(f"Error generating Excel report: {e}")
            # Fallback to CSV
            self._generate_csv_reports(screening_results, output_file)
    
    def _create_summary_sheet(self, wb, screening_results: Dict):
        """Create summary overview sheet"""
        ws = wb.create_sheet(title="Summary", index=0)
        
        # Title
        ws['A1'] = "Stock Screening Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Summary statistics
        row = 4
        ws[f'A{row}'] = "Strategy Summary"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 2
        
        # Headers
        headers = ["Strategy", "Stocks Found", "Top Stock", "Avg Score"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        row += 1
        
        # Strategy data
        for strategy_name, results_df in screening_results.items():
            if not results_df.empty:
                ws.cell(row=row, column=1, value=strategy_name)
                ws.cell(row=row, column=2, value=len(results_df))
                ws.cell(row=row, column=3, value=results_df.iloc[0]['symbol'] if 'symbol' in results_df.columns else 'N/A')
                
                avg_score = results_df['ranking_score'].mean() if 'ranking_score' in results_df.columns else 0
                ws.cell(row=row, column=4, value=f"{avg_score:.2f}")
                row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_strategy_sheet(self, wb, strategy_name: str, results_df: pd.DataFrame):
        """Create sheet for individual strategy results"""
        # Clean strategy name for sheet title
        sheet_name = strategy_name.replace('_', ' ').title()[:31]  # Excel sheet name limit
        ws = wb.create_sheet(title=sheet_name)
        
        # Add title
        ws['A1'] = f"{sheet_name} Results"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A2'] = f"Found {len(results_df)} stocks"
        
        # Select key columns for display
        display_columns = self._get_display_columns(results_df)
        display_df = results_df[display_columns].copy()
        
        # Round numeric columns
        numeric_columns = display_df.select_dtypes(include=[np.number]).columns
        display_df[numeric_columns] = display_df[numeric_columns].round(2)
        
        # Add data starting from row 4
        start_row = 4
        
        # Add headers
        for col, header in enumerate(display_df.columns, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Add data
        for row_idx, (_, row) in enumerate(display_df.iterrows(), start_row + 1):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Apply formatting
        self._format_sheet(ws, len(display_df), len(display_columns), start_row)
    
    def _create_analysis_sheet(self, wb, screening_results: Dict):
        """Create analysis and insights sheet"""
        ws = wb.create_sheet(title="Analysis")
        
        ws['A1'] = "Screening Analysis"
        ws['A1'].font = Font(size=14, bold=True)
        
        row = 3
        
        # Overall statistics
        total_stocks = sum(len(df) for df in screening_results.values())
        ws[f'A{row}'] = f"Total stocks found across all strategies: {total_stocks}"
        row += 2
        
        # Strategy performance comparison
        if len(screening_results) > 1:
            ws[f'A{row}'] = "Strategy Comparison"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            for strategy_name, results_df in screening_results.items():
                if not results_df.empty and 'ranking_score' in results_df.columns:
                    avg_score = results_df['ranking_score'].mean()
                    max_score = results_df['ranking_score'].max()
                    ws[f'A{row}'] = f"{strategy_name}: Avg Score {avg_score:.2f}, Max Score {max_score:.2f}"
                    row += 1
        
        # Common stocks across strategies
        row += 1
        ws[f'A{row}'] = "Cross-Strategy Analysis"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        if len(screening_results) > 1:
            all_symbols = []
            for results_df in screening_results.values():
                if not results_df.empty and 'symbol' in results_df.columns:
                    all_symbols.extend(results_df['symbol'].tolist())
            
            # Find stocks that appear in multiple strategies
            symbol_counts = pd.Series(all_symbols).value_counts()
            multi_strategy_stocks = symbol_counts[symbol_counts > 1]
            
            if not multi_strategy_stocks.empty:
                ws[f'A{row}'] = "Stocks appearing in multiple strategies:"
                row += 1
                for symbol, count in multi_strategy_stocks.items():
                    ws[f'A{row}'] = f"{symbol}: {count} strategies"
                    row += 1
    
    def _get_display_columns(self, df: pd.DataFrame) -> List[str]:
        """Select key columns for display in Excel"""
        priority_columns = [
            'rank', 'symbol', 'name', 'ranking_score',
            'price', 'change_rate', 'market_cap',
            'pe_ratio', 'roe', 'debt_to_equity',
            'rsi', 'price_change_1d', 'volume_ratio'
        ]
        
        # Include available priority columns
        display_columns = [col for col in priority_columns if col in df.columns]
        
        # Add a few more columns if we have space
        remaining_columns = [col for col in df.columns if col not in display_columns]
        display_columns.extend(remaining_columns[:5])  # Add up to 5 more columns
        
        return display_columns
    
    def _format_sheet(self, ws, num_rows: int, num_cols: int, start_row: int):
        """Apply formatting to worksheet"""
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(start_row, start_row + num_rows + 1):
            for col in range(1, num_cols + 1):
                ws.cell(row=row, column=col).border = thin_border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _generate_csv_reports(self, screening_results: Dict, base_filename: str):
        """Generate CSV reports as fallback when Excel is not available"""
        base_name = os.path.splitext(base_filename)[0]
        
        # Summary CSV
        summary_data = []
        for strategy_name, results_df in screening_results.items():
            if not results_df.empty:
                summary_data.append({
                    'Strategy': strategy_name,
                    'Stocks_Found': len(results_df),
                    'Top_Stock': results_df.iloc[0]['symbol'] if 'symbol' in results_df.columns else 'N/A',
                    'Avg_Score': results_df['ranking_score'].mean() if 'ranking_score' in results_df.columns else 0
                })
        
        if summary_data:
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_csv(f"{base_name}_summary.csv", index=False)
            self.logger.info(f"Summary CSV saved to {base_name}_summary.csv")
        
        # Individual strategy CSVs
        for strategy_name, results_df in screening_results.items():
            if not results_df.empty:
                csv_filename = f"{base_name}_{strategy_name}.csv"
                
                # Select display columns
                display_columns = self._get_display_columns(results_df)
                display_df = results_df[display_columns].copy()
                
                # Round numeric columns
                numeric_columns = display_df.select_dtypes(include=[np.number]).columns
                display_df[numeric_columns] = display_df[numeric_columns].round(2)
                
                display_df.to_csv(csv_filename, index=False)
                self.logger.info(f"Strategy CSV saved to {csv_filename}")
    
    def create_watchlist_export(self, results_df: pd.DataFrame, output_file: str):
        """Create a simple watchlist file with just symbols"""
        if results_df.empty or 'symbol' not in results_df.columns:
            self.logger.warning("No symbols to export for watchlist")
            return
        
        symbols = results_df['symbol'].tolist()
        
        with open(output_file, 'w') as f:
            f.write("# Stock Screening Watchlist\n")
            f.write(f"# Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"# Total symbols: {len(symbols)}\n\n")
            
            for symbol in symbols:
                f.write(f"{symbol}\n")
        
        self.logger.info(f"Watchlist exported to {output_file}")
    
    def add_performance_tracking(self, wb, screening_results: Dict):
        """Add a sheet for tracking performance over time"""
        # This would be used to track how well the screening strategies perform
        # Implementation would depend on having historical screening data
        pass
